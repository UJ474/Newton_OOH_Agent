"""
ingest_inventory.py
--------------------
Cleans, normalizes, upserts, and vectorizes all sheets from
AdOnMo Sheet.xlsx into the `utkarsh_test`
MongoDB collection.

Usage:
    MONGODB_URI="mongodb+srv://..." python ingest_inventory.py

    # Dry run (no DB writes, just prints cleaned rows):
    python ingest_inventory.py --dry-run

    # Skip vectorization (useful if you want to insert first, vectorize later):
    python ingest_inventory.py --skip-vectorize

    # Process only specific sheets:
    python ingest_inventory.py --sheets rwa coworking gym
"""

import asyncio
import argparse
import logging
import os
import re
import sys
import certifi
from pathlib import Path
from typing import Any

import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook

# ── adjust this path if running from a different working directory ──────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

print("Project root:", PROJECT_ROOT)
print("sys.path[0]:", sys.path[0])
from memory.embeddings import get_embedding_provider

# ── config ─────────────────────────────────────────────────────────────────
XLSX_PATH   = "AdOnMo Sheet.xlsx"
COLLECTION  = "utkarsh_test"

# Fields concatenated into the embedding text (must match existing index)
EMBED_FIELDS = [
    "venue_name",
    "locality",
    "city",
    "venue_category",
    "audience_tags",
    "environment_tags",
    "screen_placement_raw",
]

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def clean_str(val) -> str | None:
    """Strip whitespace, return None for empty/NA values."""
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ("", "none", "nan", "n/a", "#n/a", "na", "-"):
        return None
    return s

def clean_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None

def clean_int(val) -> int | None:
    f = clean_float(val)
    return int(f) if f is not None else None

def clean_city(val) -> str | None:
    s = clean_str(val)
    return s.strip().title() if s else None

def clean_pincode(val) -> str | None:
    s = clean_str(val)
    if not s:
        return None
    # Sometimes stored as float: 560103.0
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s

def split_lat_lng(val) -> tuple[float | None, float | None]:
    """Parse 'lat,lng' string into (lat, lng) floats."""
    s = clean_str(val)
    if not s:
        return None, None
    parts = s.split(",")
    if len(parts) >= 2:
        return clean_float(parts[0]), clean_float(parts[1])
    return None, None

def parse_footfall_str(val) -> int | None:
    """Handle footfall strings like '35L-40L', '3L-3.5L', '3L', '687704'."""
    s = clean_str(val)
    if not s:
        return None
    s_lower = s.lower().replace(",", "")
    # e.g. '35l-40l' → take average
    m = re.findall(r"(\d+\.?\d*)\s*l", s_lower)
    if m:
        nums = [float(x) * 100_000 for x in m]
        return int(sum(nums) / len(nums))
    # plain number
    f = clean_float(s)
    return int(f) if f is not None else None

def make_media_formats(val) -> list[str]:
    s = clean_str(val)
    if not s:
        return []
    # Strip wrapping parens like "(.mp4)"
    s = s.strip("()")
    return [x.strip() for x in re.split(r"[,/]", s) if x.strip()]

def make_tags(val) -> list[str]:
    """Split a comma/semicolon separated tags string into a list."""
    s = clean_str(val)
    if not s:
        return []
    return [t.strip() for t in re.split(r"[,;]", s) if t.strip()]


# ═══════════════════════════════════════════════════════════════════════════
# SHEET PARSERS  — each returns a list of clean dicts
# ═══════════════════════════════════════════════════════════════════════════

def parse_rwa(ws) -> list[dict]:
    """RWA Inventory — residential apartments."""
    rows = list(ws.iter_rows(values_only=True))
    # header is row 0
    records = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        sl, city, media_id, zone, locality, pincode, phase, profile, \
            prop_name, households, screens, avg_price, impressions, latlng = (
                row + (None,) * 14
            )[:14]

        lat, lng = split_lat_lng(latlng)
        rec = {
            "media_site_id":        clean_str(media_id),
            "venue_name":           clean_str(prop_name),
            "venue_category":       "residential",
            "city":                 clean_city(city),
            "zone":                 clean_str(zone),
            "locality":             clean_str(locality),
            "address":              None,
            "pin_code":             clean_pincode(pincode),
            "latitude":             lat,
            "longitude":            lng,
            "no_of_screens":        clean_int(screens),
            "screen_size":          None,
            "screen_type":          None,
            "pixel_ratio":          None,
            "dimensions":           None,
            "monthly_footfall":     None,
            "monthly_impressions":  clean_int(impressions),
            "audience_tags":        [clean_str(profile)] if clean_str(profile) else [],
            "environment_tags":     ["residential", "apartment"],
            "screen_placement_raw": None,
            "cost_per_month":       clean_float(avg_price),
            "slot_and_loop":        None,
            "media_formats":        [],
            "no_of_households":     clean_int(households),
            "source_dataset":       "RWA Inventory",
        }
        records.append(rec)
    log.info(f"RWA: parsed {len(records)} records")
    return records


def parse_coworking(ws) -> list[dict]:
    """Co-Working Spaces Inventory."""
    rows = list(ws.iter_rows(values_only=True))
    records = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        sno, media_id, city, zone, locality, pincode, bldg_name, \
            employees, screens, impressions, screen_size, latlng, \
            facility_type, tenant_details = (row + (None,) * 14)[:14]

        lat, lng = split_lat_lng(latlng)
        # locality sometimes has a full address stuffed in — keep first ~60 chars
        clean_loc = clean_str(locality)
        if clean_loc and len(clean_loc) > 80:
            clean_loc = clean_loc[:60] + "…"

        env_tags = ["coworking", "office"]
        if clean_str(facility_type) and "corporate" in str(facility_type).lower():
            env_tags.append("corporate")

        rec = {
            "media_site_id":        clean_str(media_id),
            "venue_name":           clean_str(bldg_name),
            "venue_category":       "coworking",
            "city":                 clean_city(city),
            "zone":                 clean_str(zone),
            "locality":             clean_loc,
            "address":              None,
            "pin_code":             clean_pincode(pincode),
            "latitude":             lat,
            "longitude":            lng,
            "no_of_screens":        clean_int(screens),
            "screen_size":          clean_str(screen_size),
            "screen_type":          "LED TV",
            "pixel_ratio":          None,
            "dimensions":           None,
            "monthly_footfall":     None,
            "monthly_impressions":  clean_int(impressions),
            "audience_tags":        make_tags(tenant_details)[:10],  # cap at 10
            "environment_tags":     env_tags,
            "screen_placement_raw": None,
            "cost_per_month":       None,
            "slot_and_loop":        None,
            "media_formats":        [],
            "no_of_employees":      clean_int(employees),
            "facility_type":        clean_str(facility_type),
            "source_dataset":       "Co-Working Spaces Inventory",
        }
        records.append(rec)
    log.info(f"Coworking: parsed {len(records)} records")
    return records


def parse_tech_park(ws) -> list[dict]:
    """Corporate Tech Park Inventory — handles merged cells by forward-filling."""
    rows = list(ws.iter_rows(values_only=True))
    records = []
    # Forward-fill state for merged-cell columns
    last_sl   = None
    last_city = None
    last_area = None
    last_park = None
    last_emp  = None
    last_re   = None
    last_net  = None
    last_ten  = None

    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        sl, city, area, park, location, emp_strength, re_allowed, \
            screen_type, dimensions, display_size, units, network, \
            key_tenants = (row + (None,) * 13)[:13]

        # Forward-fill merged columns
        if sl        is not None: last_sl   = sl
        if city      is not None: last_city = city
        if area      is not None: last_area = area
        if park      is not None: last_park = park
        if emp_strength is not None: last_emp = emp_strength
        if re_allowed   is not None: last_re  = re_allowed
        if network      is not None: last_net = network
        if key_tenants  is not None: last_ten = key_tenants

        rec = {
            "media_site_id":        None,
            "venue_name":           clean_str(last_park),
            "venue_category":       "tech_park",
            "city":                 clean_city(last_city),
            "zone":                 clean_str(last_area),
            "locality":             clean_str(last_area),
            "address":              None,
            "pin_code":             None,
            "latitude":             None,
            "longitude":            None,
            "no_of_screens":        clean_int(units),
            "screen_size":          f"{clean_str(display_size)} inches" if display_size else None,
            "screen_type":          clean_str(screen_type),
            "pixel_ratio":          None,
            "dimensions":           clean_str(dimensions),
            "monthly_footfall":     None,
            "monthly_impressions":  None,
            "audience_tags":        make_tags(last_ten)[:10] if last_ten else [],
            "environment_tags":     ["tech_park", "office", "corporate"],
            "screen_placement_raw": clean_str(location),
            "cost_per_month":       None,
            "slot_and_loop":        None,
            "media_formats":        [],
            "employee_strength":    clean_int(last_emp),
            "real_estate_allowed":  clean_str(last_re),
            "source_dataset":       "Corporate Tech Park Inventory",
        }
        records.append(rec)
    log.info(f"Tech Park: parsed {len(records)} records")
    return records


def parse_mall(ws) -> list[dict]:
    """Mall Inventory — handles merged cells for city/mall name."""
    rows = list(ws.iter_rows(values_only=True))
    records = []
    last_sl   = None
    last_city = None
    last_mall = None

    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        sl, city, mall, placement, screen_type, screens, cost, footfall, \
            screen_size, pixel_ratio, latlng, slot_loop, media_fmt, \
            creative_fmt = (row + (None,) * 14)[:14]

        if sl   is not None: last_sl   = sl
        if city is not None: last_city = city
        if mall is not None: last_mall = mall

        lat, lng = split_lat_lng(latlng)
        monthly_footfall = parse_footfall_str(footfall)

        # Determine category: mall vs cinema
        mall_name = clean_str(last_mall) or ""
        category  = "cinema" if any(
            w in mall_name.lower() for w in ("cinema", "cinemas", "film", "pvr", "inox")
        ) else "mall"

        rec = {
            "media_site_id":        None,
            "venue_name":           mall_name or None,
            "venue_category":       category,
            "city":                 clean_city(last_city),
            "zone":                 None,
            "locality":             None,
            "address":              None,
            "pin_code":             None,
            "latitude":             lat,
            "longitude":            lng,
            "no_of_screens":        clean_int(screens),
            "screen_size":          clean_str(screen_size),
            "screen_type":          clean_str(screen_type),
            "pixel_ratio":          clean_str(pixel_ratio),
            "dimensions":           None,
            "monthly_footfall":     monthly_footfall,
            "monthly_impressions":  None,
            "audience_tags":        [],
            "environment_tags":     [category, "retail", "entertainment"],
            "screen_placement_raw": clean_str(placement),
            "cost_per_month":       clean_float(cost),
            "slot_and_loop":        clean_str(slot_loop),
            "media_formats":        make_media_formats(media_fmt),
            "creative_formats":     make_media_formats(creative_fmt),
            "source_dataset":       "Mall Inventory",
        }
        records.append(rec)
    log.info(f"Mall: parsed {len(records)} records")
    return records


def parse_metro(ws) -> list[dict]:
    """Hyderabad Metro Inventory — merged station + footfall cells."""
    rows = list(ws.iter_rows(values_only=True))
    records = []
    last_station  = None
    last_footfall = None

    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        sl, station, level, sides, pixel_ratio, size, screens, \
            footfall, w, h, media_id, media_name, screens2 = (
                row + (None,) * 13
            )[:13]

        if station  is not None: last_station  = station
        if footfall is not None: last_footfall = footfall

        rec = {
            "media_site_id":        clean_str(media_id),
            "venue_name":           clean_str(last_station),
            "venue_category":       "metro",
            "city":                 "Hyderabad",
            "zone":                 None,
            "locality":             clean_str(last_station),
            "address":              None,
            "pin_code":             None,
            "latitude":             None,
            "longitude":            None,
            "no_of_screens":        clean_int(screens),
            "screen_size":          clean_str(size),
            "screen_type":          clean_str(level),
            "pixel_ratio":          clean_str(pixel_ratio),
            "dimensions":           f"{w} x {h}" if w and h else None,
            "monthly_footfall":     clean_int(last_footfall),
            "monthly_impressions":  None,
            "audience_tags":        [],
            "environment_tags":     ["metro", "transit", "commuters"],
            "screen_placement_raw": f"{clean_str(level)} - {clean_str(sides)}" if level and sides else None,
            "cost_per_month":       None,
            "slot_and_loop":        None,
            "media_formats":        [],
            "source_dataset":       "Hyderabad Metro Inventory",
        }
        records.append(rec)
    log.info(f"Metro: parsed {len(records)} records")
    return records


def parse_gym(ws) -> list[dict]:
    """Cult Gym Inventory."""
    rows = list(ws.iter_rows(values_only=True))
    records = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        sno, media_id, city, zone, locality, pincode, gym_name, \
            screens, impressions, pixel_ratio, monthly_budget, \
            screen_size, latlng = (row + (None,) * 13)[:13]

        lat, lng = split_lat_lng(latlng)
        rec = {
            "media_site_id":        clean_str(media_id),
            "venue_name":           clean_str(gym_name),
            "venue_category":       "gym",
            "city":                 clean_city(city),
            "zone":                 clean_str(zone),
            "locality":             clean_str(locality),
            "address":              None,
            "pin_code":             clean_pincode(pincode),
            "latitude":             lat,
            "longitude":            lng,
            "no_of_screens":        clean_int(screens),
            "screen_size":          clean_str(screen_size),
            "screen_type":          "LED TV",
            "pixel_ratio":          clean_str(pixel_ratio),
            "dimensions":           None,
            "monthly_footfall":     None,
            "monthly_impressions":  clean_int(impressions),
            "audience_tags":        ["fitness", "health", "gym-goers"],
            "environment_tags":     ["gym", "fitness", "health"],
            "screen_placement_raw": None,
            "cost_per_month":       clean_float(monthly_budget),
            "slot_and_loop":        None,
            "media_formats":        [],
            "source_dataset":       "Cult Gym Inventory",
        }
        records.append(rec)
    log.info(f"Gym: parsed {len(records)} records")
    return records


def parse_chai_point(ws) -> list[dict]:
    """Chai Point / Coffee Machine (office_media) Inventory."""
    rows = list(ws.iter_rows(values_only=True))
    records = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        slno, city, zone, locality, bldg_name, client_name, serial_no, \
            media_type, screens, footfall, address, pincode, lat, lng, \
            slot_loop, media_fmt, tenant_list = (row + (None,) * 17)[:17]

        rec = {
            "media_site_id":        clean_str(serial_no),
            "venue_name":           clean_str(bldg_name),
            "venue_category":       "office_media",
            "city":                 clean_city(city),
            "zone":                 clean_str(zone),
            "locality":             clean_str(locality),
            "address":              clean_str(address),
            "pin_code":             clean_pincode(pincode),
            "latitude":             clean_float(lat),
            "longitude":            clean_float(lng),
            "no_of_screens":        clean_int(screens),
            "screen_size":          None,
            "screen_type":          clean_str(media_type),
            "pixel_ratio":          None,
            "dimensions":           None,
            "monthly_footfall":     clean_int(footfall),
            "monthly_impressions":  None,
            "audience_tags":        make_tags(tenant_list) if tenant_list else [],
            "environment_tags":     ["office", "corporate", "coffee_machine"],
            "screen_placement_raw": None,
            "cost_per_month":       None,
            "slot_and_loop":        clean_str(slot_loop),
            "media_formats":        make_media_formats(media_fmt),
            "client_name":          clean_str(client_name),
            "source_dataset":       "Chai Point Inventory",
        }
        records.append(rec)
    log.info(f"Chai Point: parsed {len(records)} records")
    return records


# ═══════════════════════════════════════════════════════════════════════════
# SHEET → PARSER MAP
# ═══════════════════════════════════════════════════════════════════════════

SHEET_PARSERS = {
    "rwa":        ("RWA Inventory",                parse_rwa),
    "coworking":  ("Co-Working Spaces Inventory",  parse_coworking),
    "techpark":   ("Corporate Tech Park Inventory", parse_tech_park),
    "mall":       ("Mall Inventory",               parse_mall),
    "metro":      ("Hyderabad Metro Inventory",    parse_metro),
    "gym":        ("Cult Gym Inventory",           parse_gym),
    "chaipoint":  ("Chai Point Invent",            parse_chai_point),
}


# ═══════════════════════════════════════════════════════════════════════════
# UPSERT
# ═══════════════════════════════════════════════════════════════════════════

async def upsert_records(collection, records: list[dict]) -> int:
    """
    Upsert by media_site_id (if present) or venue_name + city + screen_placement_raw.
    Returns count of upserted docs.
    """
    upserted = 0
    for rec in records:
        if rec.get("media_site_id"):
            filter_q = {"media_site_id": rec["media_site_id"]}
        else:
            filter_q = {
                "venue_name":           rec.get("venue_name"),
                "city":                 rec.get("city"),
                "screen_placement_raw": rec.get("screen_placement_raw"),
                "source_dataset":       rec.get("source_dataset"),
            }
        result = await collection.update_one(
            filter_q,
            {"$set": rec},
            upsert=True,
        )
        if result.upserted_id or result.modified_count:
            upserted += 1
    return upserted


# ═══════════════════════════════════════════════════════════════════════════
# VECTORIZE
# ═══════════════════════════════════════════════════════════════════════════

async def vectorize_new(db, collection_name: str, provider):
    """Embed documents that don't have embedding_vector yet (reuses existing logic)."""
    collection = db[collection_name]
    cursor = collection.find({"embedding_vector": {"$exists": False}})
    docs = await cursor.to_list(length=5000)

    if not docs:
        log.info("No new documents to vectorize.")
        return

    log.info(f"Vectorizing {len(docs)} new documents…")
    ok = fail = 0
    for doc in docs:
        parts = []
        for field in EMBED_FIELDS:
            val = doc.get(field)
            if val:
                if isinstance(val, list):
                    parts.append(", ".join(str(v) for v in val))
                else:
                    parts.append(str(val))
        combined = ". ".join(parts).strip()
        if not combined:
            continue
        try:
            vectors = await provider.embed([combined])
            if vectors and vectors[0]:
                await collection.update_one(
                    {"_id": doc["_id"]},
                    {"$set": {"embedding_vector": vectors[0]}},
                )
                ok += 1
            else:
                fail += 1
        except Exception as e:
            log.error(f"Embed error for {doc['_id']}: {e}")
            fail += 1

    log.info(f"Vectorization done — success: {ok}, failed: {fail}")

# ═══════════════════════════════════════════════════════════════════════════
# EXPORT TO EXCEL
# ═══════════════════════════════════════════════════════════════════════════


# def export_cleaned_to_excel(records, filename="cleaned_inventory.xlsx"):
#     """Export cleaned records to an Excel file."""

#     if not records:
#         print("No records to export.")
#         return

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Cleaned Inventory"

#     # Get all possible columns across all records
#     headers = sorted(set().union(*(record.keys() for record in records)))
#     ws.append(headers)

#     for record in records:
#         row = []
#         for header in headers:
#             value = record.get(header)

#             # Convert lists into comma-separated strings
#             if isinstance(value, list):
#                 value = ", ".join(str(v) for v in value)

#             row.append(value)

#         ws.append(row)

#     wb.save(filename)
#     print(f"\n✅ Cleaned data exported to {filename}")

# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════



async def main(sheets_filter: list[str], dry_run: bool, skip_vectorize: bool):
    mongo_uri = os.environ.get("MONGODB_URI")
    if not mongo_uri and not dry_run:
        log.error("MONGODB_URI not set. Use --dry-run or export MONGODB_URI.")
        return

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)

    all_records: list[dict] = []

    for key, (sheet_name, parser_fn) in SHEET_PARSERS.items():
        if sheets_filter and key not in sheets_filter:
            continue
        if sheet_name not in wb.sheetnames:
            log.warning(f"Sheet '{sheet_name}' not found — skipping.")
            continue
        ws = wb[sheet_name]
        records = parser_fn(ws)
        all_records.extend(records)

    # Exporting the cleaned records to an Excel file
    # export_cleaned_to_excel(all_records)

    if dry_run:
        import json
        print("\n=== DRY RUN — first 3 records per source ===\n")
        seen = {}
        for r in all_records:
            src = r["source_dataset"]
            seen.setdefault(src, 0)
            if seen[src] < 3:
                print(json.dumps(r, indent=2, default=str))
                seen[src] += 1
        return

    # ── DB operations ────────────────────────────────────────────────────
    client = AsyncIOMotorClient(mongo_uri, tlsCAFile=certifi.where())
    db = client.get_database()
    collection = db[COLLECTION]

    log.info(f"Upserting {len(all_records)} records into '{COLLECTION}'…")
    upserted = await upsert_records(collection, all_records)
    log.info(f"Upsert complete — {upserted} docs inserted/modified.")

    if not skip_vectorize:
        provider = get_embedding_provider(db)
        await vectorize_new(db, COLLECTION, provider)

    log.info("All done.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ingest OOH inventory into MongoDB")
    parser.add_argument(
        "--sheets", nargs="*",
        choices=list(SHEET_PARSERS.keys()),
        help="Which sheets to process (default: all). "
             f"Options: {', '.join(SHEET_PARSERS.keys())}",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Parse and clean only — no DB writes, no embeddings.",
    )
    parser.add_argument(
        "--skip-vectorize", action="store_true",
        help="Insert into MongoDB but skip embedding step.",
    )
    args = parser.parse_args()
    asyncio.run(main(
        sheets_filter=args.sheets or [],
        dry_run=args.dry_run,
        skip_vectorize=args.skip_vectorize,
    ))
